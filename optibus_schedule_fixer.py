from __future__ import annotations

import argparse
import json
import zipfile
from collections import Counter, defaultdict
from copy import copy
from dataclasses import dataclass
from datetime import datetime, time
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping, Sequence, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


OUTPUT_HEADERS = [
    "Duty id",
    "Vehicle Block Id",
    "Event Type",
    "Description",
    "Trip Id",
    "Route Id",
    "Sign",
    "Direction",
    "Alternative",
    "Day Offset",
    "Start Time",
    "End Time",
    "Boarding Time",
    "Origin Stop Id",
    "Origin Stop Name",
    "Destination Stop Id",
    "Destination Stop Name",
    "Required Vehicle Type",
    "Actual Vehicle Type",
    "Distance",
    "First Stop Name",
    "Last Stop Name",
    "Overlap",
    "Days",
    "Service Groups",
    "Pref Group",
    "Service Group Days",
]

SOURCE_TO_OUTPUT = {
    "duty_id": "Duty id",
    "vehicle_block_id": "Vehicle Block Id",
    "event_type": "Event Type",
    "description": "Description",
    "trip_id": "Trip Id",
    "route_id": "Route Id",
    "sign": "Sign",
    "direction": "Direction",
    "alternative": "Alternative",
    "day_offset": "Day Offset",
    "start_time": "Start Time",
    "end_time": "End Time",
    "boarding_time": "Boarding Time",
    "origin_stop_id": "Origin Stop Id",
    "origin_stop_name": "Origin Stop Name",
    "destination_stop_id": "Destination Stop Id",
    "destination_stop_name": "Destination Stop Name",
    "required_vehicle_type": "Required Vehicle Type",
    "actual_vehicle_type": "Actual Vehicle Type",
    "distance": "Distance",
    "overlap": "Overlap",
    "days": "Days",
    "service_groups": "Service Groups",
    "pref_group": "Pref Group",
    "service_group_days": "Service Group Days",
}

COLUMN_WIDTHS = [
    12,
    17,
    15,
    14,
    12,
    28,
    12,
    12,
    20,
    12,
    12,
    12,
    14,
    16,
    34,
    18,
    34,
    20,
    18,
    12,
    20,
    20,
    12,
    12,
    18,
    14,
    18,
]


@dataclass
class FixResult:
    workbook_bytes: bytes
    stats: Dict[str, int]
    validation: Dict[str, int]


def norm_header(value: Any) -> str:
    return str(value or "").strip().lower().replace(" ", "_")


def norm_id(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def workbook_bytes_from_path(path: str | Path) -> bytes:
    return Path(path).read_bytes()


def extract_export_zip(export_zip_bytes: bytes) -> Tuple[bytes, bytes]:
    """Return (data_set_json_bytes, full_schedule_xlsx_bytes) from an Optibus export ZIP."""
    with zipfile.ZipFile(BytesIO(export_zip_bytes), "r") as zf:
        names = zf.namelist()
        data_set_names = [name for name in names if name.lower().endswith("data_set.json")]
        full_schedule_names = [
            name
            for name in names
            if name.lower().endswith(".xlsx") and "full_schedule" in Path(name).name.lower()
        ]
        if not data_set_names:
            raise ValueError("The export ZIP does not contain data_set.json.")
        if not full_schedule_names:
            raise ValueError("The export ZIP does not contain a full_schedule .xlsx file.")
        return zf.read(data_set_names[0]), zf.read(full_schedule_names[0])


def parse_ds_minutes(value: Any) -> int | None:
    if value is None:
        return None
    text = str(value)
    day = 0
    if "." in text:
        day_text, text = text.split(".", 1)
        day = int(day_text)
    hour, minute = [int(part) for part in text.split(":")[:2]]
    return day * 1440 + hour * 60 + minute


def format_minutes(value: int | None) -> str | None:
    if value is None:
        return None
    return f"{value // 60:02d}:{value % 60:02d}"


def excel_time_to_minutes(value: Any, base_date: Any = None) -> int | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        day_delta = (value.date() - base_date).days if base_date else 0
        return day_delta * 1440 + value.hour * 60 + value.minute
    if isinstance(value, time):
        return value.hour * 60 + value.minute
    text = str(value).strip()
    if not text:
        return None
    hour, minute = [int(part) for part in text.split(":")[:2]]
    return hour * 60 + minute


def json_value(value: Any, *, base_date: Any = None, is_time: bool = False) -> Any:
    if is_time:
        return format_minutes(excel_time_to_minutes(value, base_date))
    if isinstance(value, datetime):
        return format_minutes(excel_time_to_minutes(value, base_date))
    if isinstance(value, time):
        return format_minutes(value.hour * 60 + value.minute)
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def parse_output_minutes(value: Any) -> int | None:
    if value is None or value == "":
        return None
    hour, minute = [int(part) for part in str(value).split(":")[:2]]
    return hour * 60 + minute


def route_pieces(routes: Mapping[str, Dict[str, Any]], route_id: Any) -> Dict[str, Any]:
    route = routes.get(str(route_id))
    if route:
        return {
            "sign": route.get("sign"),
            "direction": route.get("direction"),
            "alternative": route.get("alternative"),
            "distance": route.get("distance"),
        }
    parts = str(route_id).split("-")
    return {
        "sign": parts[0] if len(parts) > 0 else None,
        "direction": parts[1] if len(parts) > 1 else None,
        "alternative": parts[2] if len(parts) > 2 else None,
        "distance": None,
    }


def stop_name(stops: Mapping[str, Dict[str, Any]], stop_id: Any) -> str | None:
    if stop_id is None:
        return None
    return stops.get(str(stop_id), {}).get("short_description")


def read_workbook_rows(workbook_bytes: bytes) -> Tuple[List[Tuple[Any, ...]], Dict[str, int]]:
    workbook = load_workbook(BytesIO(workbook_bytes), read_only=True, data_only=True)
    worksheet = workbook.active
    raw = list(worksheet.iter_rows(values_only=True))
    if not raw:
        raise ValueError("Workbook is empty.")
    headers = [norm_header(value) for value in raw[0]]
    return raw[1:], {name: index for index, name in enumerate(headers)}


def load_service_reference(full_schedule_bytes: bytes) -> Dict[str, Dict[str, Any]]:
    rows, index = read_workbook_rows(full_schedule_bytes)
    required = {"event_type", "trip_id", "route_id", "start_time", "end_time"}
    missing = required - set(index)
    if missing:
        raise ValueError(f"Reference full schedule is missing columns: {', '.join(sorted(missing))}")

    reference: Dict[str, Dict[str, Any]] = {}
    for row in rows:
        if row[index["event_type"]] != "service_trip":
            continue
        trip_id = norm_id(row[index["trip_id"]])
        if not trip_id:
            continue
        reference[trip_id] = {
            "Trip Id": trip_id,
            "Route Id": row[index.get("route_id")],
            "Sign": row[index.get("sign")],
            "Direction": row[index.get("direction")],
            "Alternative": row[index.get("alternative")],
            "Day Offset": row[index.get("day_offset")],
            "Start Time": row[index.get("start_time")],
            "End Time": row[index.get("end_time")],
            "Boarding Time": row[index.get("boarding_time")],
            "Origin Stop Id": row[index.get("origin_stop_id")],
            "Origin Stop Name": row[index.get("origin_stop_name")],
            "Destination Stop Id": row[index.get("destination_stop_id")],
            "Destination Stop Name": row[index.get("destination_stop_name")],
            "Required Vehicle Type": row[index.get("required_vehicle_type")],
            "Actual Vehicle Type": row[index.get("actual_vehicle_type")],
            "Distance": row[index.get("distance")],
        }
    return reference


def apply_service_reference(row_dict: Dict[str, Any], service_reference: Mapping[str, Dict[str, Any]], trip_id: Any) -> bool:
    reference = service_reference.get(str(trip_id))
    if not reference:
        return False
    row_dict.update(reference)
    return True


def shift_time_pair(row: List[Any], delta: int) -> None:
    for index in (10, 11):
        minutes = parse_output_minutes(row[index])
        if minutes is not None:
            row[index] = format_minutes(minutes + delta)


def set_end_time(row: List[Any], minutes: int) -> None:
    row[11] = format_minutes(minutes)


def remove_duty_overlaps(rows: List[List[Any]]) -> Tuple[int, int]:
    by_duty: Dict[str | None, List[List[Any]]] = defaultdict(list)
    for row in rows:
        by_duty[norm_id(row[0])].append(row)

    shifted = 0
    trimmed = 0
    for duty_rows in by_duty.values():
        duty_rows.sort(
            key=lambda row: (
                parse_output_minutes(row[10]) if parse_output_minutes(row[10]) is not None else 10**9,
                parse_output_minutes(row[11]) if parse_output_minutes(row[11]) is not None else 10**9,
                str(row[2] or ""),
                str(row[4] or ""),
            )
        )
        active: List[List[Any]] = []
        for row in duty_rows:
            start = parse_output_minutes(row[10])
            end = parse_output_minutes(row[11])
            if start is None or end is None:
                continue

            if row[2] == "service_trip":
                for active_row in active:
                    active_end = parse_output_minutes(active_row[11])
                    active_start = parse_output_minutes(active_row[10])
                    if (
                        active_row[2] != "service_trip"
                        and active_start is not None
                        and active_end is not None
                        and active_start <= start < active_end
                    ):
                        set_end_time(active_row, start)
                        trimmed += 1
                active = [
                    active_row
                    for active_row in active
                    if parse_output_minutes(active_row[11]) is not None and parse_output_minutes(active_row[11]) > start
                ]
            elif active:
                max_active_end = max(
                    parse_output_minutes(active_row[11])
                    for active_row in active
                    if parse_output_minutes(active_row[11]) is not None
                )
                if start < max_active_end:
                    delta = max_active_end - start
                    shift_time_pair(row, delta)
                    shifted += 1
                    start += delta
                    end += delta

            active = [
                active_row
                for active_row in active
                if parse_output_minutes(active_row[11]) is not None and parse_output_minutes(active_row[11]) > start
            ]
            active.append(row)
    return shifted, trimmed


def split_overlapping_duty_ids(rows: List[List[Any]]) -> Tuple[int, int]:
    by_duty: Dict[str, List[List[Any]]] = defaultdict(list)
    for row in rows:
        duty = norm_id(row[0])
        if duty:
            by_duty[duty].append(row)

    remapped = 0
    split_duties = 0
    for duty, duty_rows in by_duty.items():
        lanes: List[int] = []
        duty_rows.sort(
            key=lambda row: (
                parse_output_minutes(row[10]) if parse_output_minutes(row[10]) is not None else 10**9,
                parse_output_minutes(row[11]) if parse_output_minutes(row[11]) is not None else 10**9,
                str(row[2] or ""),
                str(row[4] or ""),
            )
        )
        for row in duty_rows:
            start = parse_output_minutes(row[10])
            end = parse_output_minutes(row[11])
            if start is None or end is None:
                continue
            lane = next((index for index, lane_end in enumerate(lanes) if start >= lane_end), None)
            if lane is None:
                lane = len(lanes)
                lanes.append(-1)
            lanes[lane] = max(lanes[lane], end)
            if lane > 0:
                row[0] = f"{duty}_{lane + 1}"
                remapped += 1
        if len(lanes) > 1:
            split_duties += 1
    return remapped, split_duties


def split_overlapping_vehicle_blocks(rows: List[List[Any]]) -> Tuple[int, int]:
    by_vehicle: Dict[str, List[List[Any]]] = defaultdict(list)
    for row in rows:
        vehicle = norm_id(row[1])
        if vehicle:
            by_vehicle[vehicle].append(row)

    remapped = 0
    split_vehicles = 0
    for vehicle, vehicle_rows in by_vehicle.items():
        lanes: List[int] = []
        lane_by_duty: Dict[str, int] = {}
        vehicle_rows.sort(
            key=lambda row: (
                parse_output_minutes(row[10]) if parse_output_minutes(row[10]) is not None else 10**9,
                parse_output_minutes(row[11]) if parse_output_minutes(row[11]) is not None else 10**9,
                str(row[0] or ""),
                str(row[4] or ""),
            )
        )
        for row in vehicle_rows:
            start = parse_output_minutes(row[10])
            end = parse_output_minutes(row[11])
            if start is None or end is None:
                continue
            duty = norm_id(row[0])
            lane = lane_by_duty.get(duty) if duty else None
            if lane is not None and start < lanes[lane]:
                lane = None
            if lane is None:
                lane = next((index for index, lane_end in enumerate(lanes) if start >= lane_end), None)
            if lane is None:
                lane = len(lanes)
                lanes.append(-1)
            if duty:
                lane_by_duty[duty] = lane
            lanes[lane] = max(lanes[lane], end)
            if lane > 0:
                row[1] = f"{vehicle}_{lane + 1}"
                remapped += 1
        if len(lanes) > 1:
            split_vehicles += 1
    return remapped, split_vehicles


def remove_duplicate_events(rows: List[List[Any]]) -> Tuple[List[List[Any]], int]:
    seen = set()
    deduped = []
    removed = 0
    for row in rows:
        container = norm_id(row[1]) if norm_id(row[1]) else f"duty:{norm_id(row[0])}"
        key = (container, row[2], row[4], row[5], row[10], row[11], row[13], row[15])
        if key in seen:
            removed += 1
            continue
        seen.add(key)
        deduped.append(row)
    return deduped, removed


def sort_rows(rows: List[List[Any]]) -> List[List[Any]]:
    return sorted(
        rows,
        key=lambda row: (
            str(row[0] or ""),
            parse_output_minutes(row[10]) if parse_output_minutes(row[10]) is not None else 10**9,
            parse_output_minutes(row[11]) if parse_output_minutes(row[11]) is not None else 10**9,
            str(row[1] or ""),
            str(row[4] or ""),
            str(row[2] or ""),
        ),
    )


def write_schedule_workbook(rows: Sequence[Sequence[Any]]) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet.append(OUTPUT_HEADERS)
    for row in rows:
        worksheet.append(list(row))

    fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
    for index, width in enumerate(COLUMN_WIDTHS, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_data_set_indexes(data_set: Dict[str, Any]) -> Tuple[Dict[str, Dict[str, Any]], Dict[str, Dict[str, Any]], Dict[str, Dict[str, Any]]]:
    routes = {str(route["id"]): route for route in data_set.get("routes", [])}
    stops = {str(stop["id"]): stop for stop in data_set.get("stops", [])}
    trips = {}
    for trip in data_set.get("trips", []):
        for key in (trip.get("id"), trip.get("user_id"), trip.get("trip_code")):
            if key is not None:
                trips[str(key)] = trip
    return routes, stops, trips


def build_schedule_rows(schedule_bytes: bytes, data_set: Dict[str, Any], service_reference: Mapping[str, Dict[str, Any]]) -> Tuple[List[List[Any]], Dict[str, int]]:
    routes, stops, trips = build_data_set_indexes(data_set)
    raw_rows, source_index = read_workbook_rows(schedule_bytes)
    required = {"event_type", "start_time", "end_time"}
    missing = required - set(source_index)
    if missing:
        raise ValueError(f"Schedule workbook is missing columns: {', '.join(sorted(missing))}")

    base_dates = [
        value.date()
        for row in raw_rows
        for value in (row[source_index["start_time"]], row[source_index["end_time"]])
        if isinstance(value, datetime)
    ]
    base_date = min(base_dates) if base_dates else None

    rows: List[List[Any]] = []
    expanded_parent_rows = 0
    for raw in raw_rows:
        output = {header: None for header in OUTPUT_HEADERS}
        for source_name, output_name in SOURCE_TO_OUTPUT.items():
            if source_name not in source_index:
                continue
            output[output_name] = json_value(
                raw[source_index[source_name]],
                base_date=base_date,
                is_time=source_name in {"start_time", "end_time"},
            )

        if output["Event Type"] == "service_trip":
            trip_id = norm_id(output["Trip Id"])
            trip = trips.get(trip_id or "")
            sub_trips = trip.get("sub_trips", []) if trip else []
            if len(sub_trips) > 1:
                expanded_parent_rows += 1
                for sub_trip in sub_trips:
                    child = copy(output)
                    child["Trip Id"] = sub_trip.get("id")
                    if not apply_service_reference(child, service_reference, sub_trip.get("id")):
                        child["Route Id"] = sub_trip.get("route_id")
                        route = route_pieces(routes, sub_trip.get("route_id"))
                        child["Sign"] = route["sign"]
                        child["Direction"] = route["direction"]
                        child["Alternative"] = route["alternative"]
                        child["Start Time"] = format_minutes(parse_ds_minutes(sub_trip.get("departure_time")))
                        child["End Time"] = format_minutes(parse_ds_minutes(sub_trip.get("arrival_time")))
                        child["Origin Stop Id"] = sub_trip.get("origin_stop_id")
                        child["Origin Stop Name"] = stop_name(stops, sub_trip.get("origin_stop_id"))
                        child["Destination Stop Id"] = sub_trip.get("destination_stop_id")
                        child["Destination Stop Name"] = stop_name(stops, sub_trip.get("destination_stop_id"))
                        child["Distance"] = route["distance"]
                    rows.append([child[header] for header in OUTPUT_HEADERS])
                continue
            apply_service_reference(output, service_reference, trip_id)

        rows.append([output[header] for header in OUTPUT_HEADERS])

    duty_rows_shifted, duty_rows_trimmed = remove_duty_overlaps(rows)
    duty_rows_remapped, duties_split = split_overlapping_duty_ids(rows)
    vehicle_rows_remapped, vehicles_split = split_overlapping_vehicle_blocks(rows)
    rows, duplicate_rows_removed = remove_duplicate_events(rows)

    stats = {
        "input_rows": len(raw_rows),
        "output_rows": len(rows),
        "expanded_parent_rows": expanded_parent_rows,
        "duty_rows_shifted": duty_rows_shifted,
        "duty_rows_trimmed": duty_rows_trimmed,
        "duty_rows_remapped": duty_rows_remapped,
        "duties_split": duties_split,
        "vehicle_rows_remapped": vehicle_rows_remapped,
        "vehicles_split": vehicles_split,
        "duplicate_rows_removed": duplicate_rows_removed,
    }
    return sort_rows(rows), stats


def validation_indexes(data_set: Dict[str, Any]) -> Tuple[set[str], Dict[str, int], Dict[str, str]]:
    accepted = set()
    parent_counts = {}
    sub_to_parent = {}
    for trip in data_set.get("trips", []):
        for key in (trip.get("id"), trip.get("user_id"), trip.get("trip_code")):
            if key is not None:
                accepted.add(str(key))
        subs = [str(sub_trip.get("id")) for sub_trip in trip.get("sub_trips", []) if sub_trip.get("id") is not None]
        for sub_id in subs:
            accepted.add(sub_id)
        if len(subs) > 1:
            parent = str(trip.get("user_id") or trip.get("trip_code") or trip.get("id"))
            parent_counts[parent] = len(subs)
            for sub_id in subs:
                sub_to_parent[sub_id] = parent
    return accepted, parent_counts, sub_to_parent


def count_overlaps(rows: Sequence[Sequence[Any]], group_index: int) -> int:
    groups: Dict[str, List[Sequence[Any]]] = defaultdict(list)
    for row in rows:
        key = norm_id(row[group_index])
        if key:
            groups[key].append(row)
    overlap_count = 0
    for items in groups.values():
        sorted_items = sorted(
            items,
            key=lambda row: (
                parse_output_minutes(row[10]) if parse_output_minutes(row[10]) is not None else 10**9,
                parse_output_minutes(row[11]) if parse_output_minutes(row[11]) is not None else 10**9,
            ),
        )
        active: List[Sequence[Any]] = []
        for item in sorted_items:
            start = parse_output_minutes(item[10])
            if start is None:
                continue
            for previous in active:
                previous_end = parse_output_minutes(previous[11])
                if previous_end is not None and start < previous_end:
                    overlap_count += 1
            active = [
                previous
                for previous in active
                if parse_output_minutes(previous[11]) is not None and parse_output_minutes(previous[11]) > start
            ]
            active.append(item)
    return overlap_count


def validate_rows(rows: Sequence[Sequence[Any]], data_set: Dict[str, Any]) -> Dict[str, int]:
    accepted, parent_counts, sub_to_parent = validation_indexes(data_set)
    event_counts = Counter(row[2] for row in rows)
    service_ids = [norm_id(row[4]) for row in rows if row[2] == "service_trip"]
    missing_ids = [trip_id for trip_id in service_ids if trip_id not in accepted]
    parent_ids = [trip_id for trip_id in service_ids if trip_id in parent_counts]

    by_vehicle_parent: Dict[Tuple[str | None, str], set[str]] = defaultdict(set)
    by_duty_parent: Dict[Tuple[str | None, str], set[str]] = defaultdict(set)
    for row in rows:
        if row[2] != "service_trip":
            continue
        trip_id = norm_id(row[4])
        parent = sub_to_parent.get(trip_id or "")
        if not parent:
            continue
        by_vehicle_parent[(norm_id(row[1]), parent)].add(trip_id or "")
        by_duty_parent[(norm_id(row[0]), parent)].add(trip_id or "")

    incomplete_vehicle = sum(1 for (_, parent), subs in by_vehicle_parent.items() if len(subs) != parent_counts[parent])
    incomplete_duty = sum(1 for (_, parent), subs in by_duty_parent.items() if len(subs) != parent_counts[parent])

    signatures = Counter(
        (
            norm_id(row[1]) or f"duty:{norm_id(row[0])}",
            row[2],
            norm_id(row[4]),
            norm_id(row[5]),
            row[10],
            row[11],
            norm_id(row[13]),
            norm_id(row[15]),
        )
        for row in rows
    )
    duplicate_events = sum(count - 1 for count in signatures.values() if count > 1)

    return {
        "service_rows": event_counts.get("service_trip", 0),
        "deadhead_rows": event_counts.get("deadhead", 0),
        "sign_on_rows": event_counts.get("sign_on", 0),
        "sign_off_rows": event_counts.get("sign_off", 0),
        "depot_pull_out_rows": event_counts.get("depot_pull_out", 0),
        "depot_pull_in_rows": event_counts.get("depot_pull_in", 0),
        "duplicate_service_ids": len(service_ids) - len(set(service_ids)),
        "missing_service_ids": len(missing_ids),
        "parent_ids_still_present": len(parent_ids),
        "incomplete_subtrips_by_vehicle": incomplete_vehicle,
        "incomplete_subtrips_by_duty": incomplete_duty,
        "duplicate_events": duplicate_events,
        "duty_overlaps": count_overlaps(rows, 0),
        "vehicle_overlaps": count_overlaps(rows, 1),
    }


def fix_schedule(
    schedule_bytes: bytes,
    *,
    data_set_json_bytes: bytes | None = None,
    full_schedule_bytes: bytes | None = None,
    export_zip_bytes: bytes | None = None,
) -> FixResult:
    if export_zip_bytes:
        data_set_json_bytes, full_schedule_bytes = extract_export_zip(export_zip_bytes)
    if data_set_json_bytes is None:
        raise ValueError("Upload data_set.json or a full Optibus export ZIP containing data_set.json.")
    if full_schedule_bytes is None:
        raise ValueError("Upload a reference full schedule .xlsx or a full Optibus export ZIP containing one.")

    data_set = json.loads(data_set_json_bytes.decode("utf-8-sig"))
    service_reference = load_service_reference(full_schedule_bytes)
    rows, stats = build_schedule_rows(schedule_bytes, data_set, service_reference)
    validation = validate_rows(rows, data_set)
    return FixResult(write_schedule_workbook(rows), stats, validation)


def fix_schedule_files(
    schedule_path: str | Path,
    output_path: str | Path,
    *,
    export_zip_path: str | Path | None = None,
    data_set_json_path: str | Path | None = None,
    full_schedule_path: str | Path | None = None,
) -> FixResult:
    result = fix_schedule(
        workbook_bytes_from_path(schedule_path),
        export_zip_bytes=workbook_bytes_from_path(export_zip_path) if export_zip_path else None,
        data_set_json_bytes=workbook_bytes_from_path(data_set_json_path) if data_set_json_path else None,
        full_schedule_bytes=workbook_bytes_from_path(full_schedule_path) if full_schedule_path else None,
    )
    Path(output_path).write_bytes(result.workbook_bytes)
    return result


def main() -> None:
    parser = argparse.ArgumentParser(description="Fix an Optibus full/vehicle/crew schedule workbook for import.")
    parser.add_argument("--schedule", required=True, help="Schedule workbook to fix (.xlsx/.xls).")
    parser.add_argument("--out", required=True, help="Output .xlsx path.")
    parser.add_argument("--export-zip", help="Full Optibus export ZIP containing data_set.json and export_full_schedule.xlsx.")
    parser.add_argument("--data-set-json", help="data_set.json path, used when --export-zip is not provided.")
    parser.add_argument("--full-schedule", help="Reference full schedule .xlsx path, used when --export-zip is not provided.")
    args = parser.parse_args()

    result = fix_schedule_files(
        args.schedule,
        args.out,
        export_zip_path=args.export_zip,
        data_set_json_path=args.data_set_json,
        full_schedule_path=args.full_schedule,
    )
    print(json.dumps({"stats": result.stats, "validation": result.validation}, indent=2))


if __name__ == "__main__":
    main()
